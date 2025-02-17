# import pandas as pd
# import spacy
# from sentence_transformers import SentenceTransformer, util
# import streamlit as st
# from openpyxl import load_workbook
# import re

# # Load NLP Model
# nlp = spacy.load("en_core_web_sm")  # Named Entity Recognition
# embedder = SentenceTransformer("all-MiniLM-L6-v2")  # Sentence embeddings

# # Streamlit App UI
# def main():
#     st.title("Language Analysis of Requirements")
#     glossary_file = st.file_uploader("Upload Glossary File (Excel)", type=["xlsx"])
#     req_file = st.file_uploader("Upload Requirements File (Excel)", type=["xlsm"])
    
#     if glossary_file and req_file:
#         # Load the workbook for extracting kept and deleted terms
#         wb = load_workbook(glossary_file, data_only=True)
#         ws = wb["Notary Glossary of Terms"]  # Load "Notary Glossary of Terms" sheet

#         # Lists to store kept and deleted terms
#         kept_terms = []
#         deleted_terms = []

#         # Iterate through the column containing terms (Column A)
#         for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):  # Column A (1st Column)
#             for cell in row:
#                 term = cell.value
#                 if term:  # Ensure it's not empty
#                     # Check if the text is struck through
#                     if cell.font and cell.font.strike:
#                         deleted_terms.append(term.strip())
#                     else:
#                         kept_terms.append(term.strip())

#         # Load Requirement Data
#         def load_requirements(req_file, sheet_name, column_index):
#             df = pd.read_excel(req_file, sheet_name=sheet_name, engine="openpyxl", header=1)
#             column_name = df.columns[column_index]  # Fetch column name dynamically
#             return df[column_name].dropna().tolist()

#         functional_reqs = load_requirements(req_file, "Final Approved - Functional", 8)  # Column I (index 8)
#         nonfunctional_reqs = load_requirements(req_file, "Non-Functional", 6)  # Column G (index 6)
        
#         st.write("### Processing... This may take a moment")
        
#         # Match Terms using NLP & Exact Case-Sensitive Whole-Word Matching with Variations
#         def find_terms(glossary_terms, descriptions):
#             term_counts = {term: 0 for term in glossary_terms}
#             for desc in descriptions:
#                 for term in glossary_terms:
#                     # Perform strict case-sensitive whole-word matching, allowing minor variations
#                     base_term = re.escape(term)
#                     pattern = rf"\b{base_term}(ed|s|ing)?\b"  # Handle minor variations like plural and past tense
#                     matches = re.findall(pattern, desc)
#                     term_counts[term] += len(matches)
#             return term_counts

#         func_kept_counts = find_terms(kept_terms, functional_reqs)
#         func_deleted_counts = find_terms(deleted_terms, functional_reqs)
#         nonfunc_kept_counts = find_terms(kept_terms, nonfunctional_reqs)
#         nonfunc_deleted_counts = find_terms(deleted_terms, nonfunctional_reqs)
        
#         # Convert to DataFrame for display
#         df_func_kept = pd.DataFrame(func_kept_counts.items(), columns=["Functional Kept Term", "Count"])
#         df_func_deleted = pd.DataFrame(func_deleted_counts.items(), columns=["Functional Deleted Term", "Count"])
#         df_nonfunc_kept = pd.DataFrame(nonfunc_kept_counts.items(), columns=["Non-Functional Kept Term", "Count"])
#         df_nonfunc_deleted = pd.DataFrame(nonfunc_deleted_counts.items(), columns=["Non-Functional Deleted Term", "Count"])
        
#         # Display Results
#         st.write("### Functional Kept Terms")
#         st.dataframe(df_func_kept)
#         st.write("### Functional Deleted Terms")
#         st.dataframe(df_func_deleted)
#         st.write("### Non-Functional Kept Terms")
#         st.dataframe(df_nonfunc_kept)
#         st.write("### Non-Functional Deleted Terms")
#         st.dataframe(df_nonfunc_deleted)
        
# if __name__ == "__main__":
#     main()

import pandas as pd
import spacy
from sentence_transformers import SentenceTransformer, util
import streamlit as st
from openpyxl import load_workbook
import re

# Load NLP Model
nlp = spacy.load("en_core_web_sm")  # Named Entity Recognition
embedder = SentenceTransformer("all-MiniLM-L6-v2")  # Sentence embeddings

# Streamlit App UI
def main():
    st.title("Language Analysis of Requirements")
    glossary_file = st.file_uploader("Upload Glossary File (Excel)", type=["xlsx"])
    req_file = st.file_uploader("Upload Requirements File (Excel)", type=["xlsm"])
    
    if glossary_file and req_file:
        # Load the workbook for extracting kept and deleted terms
        wb = load_workbook(glossary_file, data_only=True)
        ws = wb["Notary Glossary of Terms"]  # Load "Notary Glossary of Terms" sheet

        # Lists to store kept and deleted terms
        kept_terms = []
        deleted_terms = []

        # Iterate through the column containing terms (Column A)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):  # Column A (1st Column)
            for cell in row:
                term = cell.value
                if term:  # Ensure it's not empty
                    # Check if the text is struck through
                    if cell.font and cell.font.strike:
                        deleted_terms.append(term.strip())
                    else:
                        kept_terms.append(term.strip())

        # Load Requirement Data
        def load_requirements(req_file, sheet_name, column_index):
            df = pd.read_excel(req_file, sheet_name=sheet_name, engine="openpyxl", header=1)
            column_name = df.columns[column_index]  # Fetch column name dynamically
            return df[column_name].dropna().tolist()

        functional_reqs = load_requirements(req_file, "Final Approved - Functional", 8)  # Column I (index 8)
        nonfunctional_reqs = load_requirements(req_file, "Non-Functional", 6)  # Column G (index 6)
        
        st.write("### Processing... This may take a moment")
        
        # Match Terms using NLP & Exact Case-Sensitive Whole-Word Matching with Variations
        def find_terms(glossary_terms, descriptions):
            term_counts = {term: 0 for term in glossary_terms}
            for desc in descriptions:
                for term in glossary_terms:
                    # Perform strict case-sensitive whole-word matching, allowing minor variations
                    base_term = re.escape(term)
                    pattern = rf"\b{base_term}(ed|s|ing)?\b"  # Handle minor variations like plural and past tense
                    matches = re.findall(pattern, desc)
                    term_counts[term] += len(matches)
            return term_counts

        func_kept_counts = find_terms(kept_terms, functional_reqs)
        func_deleted_counts = find_terms(deleted_terms, functional_reqs)
        nonfunc_kept_counts = find_terms(kept_terms, nonfunctional_reqs)
        nonfunc_deleted_counts = find_terms(deleted_terms, nonfunctional_reqs)
        
        # Ensure all lists are of the same length by padding with empty strings or zeros
        max_length = max(len(func_kept_counts), len(func_deleted_counts), len(nonfunc_kept_counts), len(nonfunc_deleted_counts))
        
        def pad_dict(d, max_length):
            return list(d.keys()) + [""] * (max_length - len(d)), list(d.values()) + [0] * (max_length - len(d))
        
        fk_terms, fk_counts = pad_dict(func_kept_counts, max_length)
        fd_terms, fd_counts = pad_dict(func_deleted_counts, max_length)
        nk_terms, nk_counts = pad_dict(nonfunc_kept_counts, max_length)
        nd_terms, nd_counts = pad_dict(nonfunc_deleted_counts, max_length)
        
        # Combine results into a single DataFrame for better visualization
        df_results = pd.DataFrame({
            "Functional Kept Term": fk_terms,
            "Functional Kept Count": fk_counts,
            "Functional Deleted Term": fd_terms,
            "Functional Deleted Count": fd_counts,
            "Non-Functional Kept Term": nk_terms,
            "Non-Functional Kept Count": nk_counts,
            "Non-Functional Deleted Term": nd_terms,
            "Non-Functional Deleted Count": nd_counts
        })
        
        # Display Results in a single dataframe
        st.write("### Combined Analysis Results")
        st.dataframe(df_results)
        
if __name__ == "__main__":
    main()
